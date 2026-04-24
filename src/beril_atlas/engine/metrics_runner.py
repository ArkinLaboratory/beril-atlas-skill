"""
Atlas L4 metrics runner — exports the registered views as CSV and XLSX.

Phase 2a deliverable. Reads from a Phase 1 warehouse (atlas.duckdb), runs
every view in beril_atlas.engine.metrics.VIEWS, writes:
  - One CSV per view at <outputs>/metrics/csv/<view_name>.csv
  - One multi-sheet XLSX at <outputs>/metrics/atlas_metrics.xlsx
    with one sheet per view + a Provenance sheet listing all view definitions
  - A summary JSON at <outputs>/metrics/run_summary.json

Usage (from spike/beril-extended/):
    python -m beril_atlas.engine.metrics_runner \\
        --warehouse ~/.beril-atlas/runs/<ts>/atlas.duckdb \\
        --outputs   ~/.beril-atlas/runs/<ts>/

Slash-command (planned):
    /beril-atlas report  # invokes this after a fresh scan

Phase 2b will add LLM-derived metric views to the same registry.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path


import duckdb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from . import metrics as m


def parse_args(argv: list[str] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="BERIL Atlas L4 metrics runner — exports CSV + XLSX",
    )
    parser.add_argument("--warehouse", type=Path, required=True,
                        help="Path to atlas.duckdb produced by `beril-atlas scan`")
    parser.add_argument("--outputs", type=Path, required=True,
                        help="Root for metrics output (e.g., ~/.beril-atlas/runs/<ts>/)")
    parser.add_argument("--quiet", action="store_true")
    return parser.parse_args(argv)


def _log(msg: str, quiet: bool = False) -> None:
    if not quiet:
        print(f"[atlas-metrics] {msg}", flush=True)


def export_csvs(results: list[m.ViewResult], csv_dir: Path) -> list[Path]:
    """Write one CSV per view; return list of paths created."""
    csv_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for r in results:
        if r.error is not None:
            continue
        path = csv_dir / f"{r.view.name}.csv"
        # DuckDB returns LIST columns as Python lists; join for CSV readability
        df = r.df.copy()
        for col in df.columns:
            if df[col].dtype == "object":
                df[col] = df[col].apply(
                    lambda v: "; ".join(map(str, v)) if isinstance(v, list) else v
                )
        df.to_csv(path, index=False)
        written.append(path)
    return written


def _cellify(val):
    """Coerce arbitrary pandas/numpy/Python values into something openpyxl accepts.

    Handles: pandas NA/NaN/NaT → None; numpy arrays / Python lists → '; '.join;
    dates → ISO string; numpy scalars → Python scalars.
    """
    import numpy as np
    import pandas as pd

    # Sequence types (numpy array, list, tuple) — flatten to string
    if isinstance(val, (list, tuple, np.ndarray)):
        return "; ".join(_cellify(x) if not isinstance(x, str) else x for x in val
                          if x is not None) if len(val) > 0 else ""

    # NA / NaN / NaT detection (pd.isna is safe on scalars)
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass

    # Date / datetime → ISO
    if isinstance(val, (dt.date, dt.datetime)):
        return val.isoformat()

    # Numpy scalar → Python scalar
    if isinstance(val, np.generic):
        return val.item()

    return val


def _set_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2E5984", end_color="2E5984", fill_type="solid")
    cell.alignment = Alignment(vertical="center")


def _set_provenance_style(cell) -> None:
    cell.font = Font(italic=True, color="555555")


def export_xlsx(results: list[m.ViewResult], xlsx_path: Path) -> Path:
    """Write a multi-sheet XLSX. One sheet per view + a TOC + a provenance sheet."""
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Table of Contents ---
    toc = wb.active
    toc.title = "TOC"
    toc.append(["BERIL Atlas — Phase 2a Metrics Export"])
    toc.append([])
    toc.append(["sheet_name", "title", "category", "headline", "rows"])
    _set_header_style(toc.cell(row=3, column=1))
    _set_header_style(toc.cell(row=3, column=2))
    _set_header_style(toc.cell(row=3, column=3))
    _set_header_style(toc.cell(row=3, column=4))
    _set_header_style(toc.cell(row=3, column=5))

    for r in results:
        if r.error is not None:
            toc.append([r.view.name, r.view.title, r.view.category, "", f"ERROR: {r.error}"])
            continue
        toc.append([
            r.view.name,
            r.view.title,
            r.view.category,
            "★" if r.view.is_headline else "",
            r.row_count,
        ])

    # --- Sheets per view ---
    for r in results:
        if r.error is not None:
            continue
        # Sheet names limited to 31 chars by Excel
        sheet_name = r.view.name[:31]
        ws = wb.create_sheet(sheet_name)

        # Title row
        ws.cell(row=1, column=1, value=r.view.title).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=r.view.description).font = Font(italic=True, color="555555")
        ws.cell(row=3, column=1, value=f"Category: {r.view.category}    Rows: {r.row_count}    "
                                      f"{'★ HEADLINE' if r.view.is_headline else ''}")

        # Header row
        header_row = 5
        for col_idx, col_name in enumerate(r.df.columns, start=1):
            c = ws.cell(row=header_row, column=col_idx, value=col_name)
            _set_header_style(c)

        # Data rows
        import numpy as np
        import pandas as pd
        for row_idx, row in enumerate(r.df.itertuples(index=False), start=header_row + 1):
            for col_idx, val in enumerate(row, start=1):
                val = _cellify(val)
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-fit-ish column widths (cap at 60)
        from openpyxl.utils import get_column_letter
        for col_idx, col_name in enumerate(r.df.columns, start=1):
            try:
                widths = [len(str(_cellify(row[col_idx - 1]) or ""))
                          for row in r.df.itertuples(index=False)]
            except Exception:
                widths = []
            max_width = max([len(str(col_name)), 12, *widths]) if widths else max(len(str(col_name)), 12)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 60)

    # --- Provenance sheet (column definitions per view) ---
    prov = wb.create_sheet("__provenance__")
    prov.append(["view_name", "column_name", "definition"])
    for col_idx in range(1, 4):
        _set_header_style(prov.cell(row=1, column=col_idx))
    for r in results:
        for col_name, defn in r.view.columns.items():
            prov.append([r.view.name, col_name, defn])

    wb.save(xlsx_path)
    return xlsx_path


def write_run_summary(results: list[m.ViewResult],
                       outputs: Path,
                       warehouse_path: Path,
                       csv_paths: list[Path],
                       xlsx_path: Path) -> Path:
    """Write a JSON summary of the metrics run."""
    summary_path = outputs / "metrics" / "run_summary.json"
    summary = {
        "ran_at": dt.datetime.utcnow().isoformat(),
        "warehouse": str(warehouse_path),
        "views_total": len(results),
        "views_succeeded": sum(1 for r in results if r.error is None),
        "views_failed": sum(1 for r in results if r.error is not None),
        "views_with_sanity_warnings": sum(1 for r in results
                                          if r.error is None and not r.sanity_passed),
        "csv_files": [str(p) for p in csv_paths],
        "xlsx_file": str(xlsx_path),
        "per_view": [
            {
                "name": r.view.name,
                "title": r.view.title,
                "category": r.view.category,
                "is_headline": r.view.is_headline,
                "row_count": r.row_count,
                "sanity_passed": r.sanity_passed,
                "expected_min_rows": r.view.expected_min_rows,
                "error": r.error,
            }
            for r in results
        ],
    }
    summary_path.write_text(json.dumps(summary, indent=2, default=str))
    return summary_path


def main(argv: list[str] = None) -> int:
    args = parse_args(argv)
    quiet = args.quiet

    if not args.warehouse.exists():
        _log(f"FATAL: warehouse not found: {args.warehouse}", quiet=False)
        return 1

    _log(f"warehouse: {args.warehouse}", quiet)
    _log(f"outputs:   {args.outputs}/metrics/", quiet)
    _log(f"running {len(m.VIEWS)} views...", quiet)

    con = duckdb.connect(str(args.warehouse), read_only=True)
    results = m.run_all_views(con)
    con.close()

    failed = [r for r in results if r.error is not None]
    sanity_warnings = [r for r in results if r.error is None and not r.sanity_passed]

    metrics_dir = args.outputs / "metrics"
    csv_dir = metrics_dir / "csv"
    xlsx_path = metrics_dir / "atlas_metrics.xlsx"

    csv_paths = export_csvs(results, csv_dir)
    export_xlsx(results, xlsx_path)
    summary_path = write_run_summary(results, args.outputs, args.warehouse, csv_paths, xlsx_path)

    _log(f"  exported {len(csv_paths)} CSVs, 1 XLSX, 1 summary JSON", quiet)
    _log(f"  views OK: {len(results) - len(failed)}/{len(results)}", quiet)
    if sanity_warnings:
        _log(f"  WARNING: {len(sanity_warnings)} views below expected_min_rows:", quiet=False)
        for r in sanity_warnings:
            _log(f"    {r.view.name}: got {r.row_count}, expected ≥{r.view.expected_min_rows}", quiet=False)
    if failed:
        _log(f"  ERROR: {len(failed)} views FAILED:", quiet=False)
        for r in failed:
            _log(f"    {r.view.name}: {r.error}", quiet=False)
        return 1

    _log(f"PASS — metrics written to {metrics_dir}", quiet)
    _log(f"  XLSX: {xlsx_path}", quiet)
    _log(f"  summary: {summary_path}", quiet)
    return 0


if __name__ == "__main__":
    sys.exit(main())
