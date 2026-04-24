"""
Tests for atlas_lib.metrics and atlas_metrics (L4 runner).

Exercises every registered view against a Phase 1 warehouse built from the
real corpus. Also verifies the CSV + XLSX exporters produce expected files.

Run from spike/beril-extended/:
    TMPDIR=/tmp/pytest-atlas python -m pytest tests/atlas/test_metrics.py -v
"""

from __future__ import annotations

import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent.parent / "scripts"
sys.path.insert(0, str(SCRIPTS))

import duckdb
import pytest

from beril_atlas.engine import metrics_runner as atlas_metrics
from beril_atlas.engine import scan as atlas_scan
from beril_atlas.engine import metrics as m


CORPUS = HERE.parent.parent / "projects"


# --------------------------------------------------------------------------
# Shared warehouse fixture — one scan, reused across all tests
# --------------------------------------------------------------------------

@pytest.fixture(scope="module")
def real_warehouse(tmp_path_factory):
    if not CORPUS.exists():
        pytest.skip("BERIL corpus not available")
    outputs = tmp_path_factory.mktemp("atlas_metrics_test")
    rc = atlas_scan.main([
        "--projects-root", str(CORPUS),
        "--outputs-root", str(outputs),
        "--quiet",
    ])
    assert rc == 0, "Phase 1 scanner failed; cannot test Phase 2a"
    return outputs / "atlas.duckdb", outputs


# --------------------------------------------------------------------------
# View registry integrity
# --------------------------------------------------------------------------

class TestViewRegistry:

    def test_all_views_have_names(self):
        for v in m.VIEWS:
            assert v.name, f"View without name: {v.title}"
            assert "." not in v.name, f"Name contains dot (CSV-unsafe): {v.name}"

    def test_all_view_names_unique(self):
        names = [v.name for v in m.VIEWS]
        assert len(names) == len(set(names)), f"Duplicate view names: {names}"

    def test_headline_views_documented(self):
        # At least half the headline-12 should have is_headline=True
        headline_count = sum(1 for v in m.VIEWS if v.is_headline)
        assert headline_count >= 8, f"Expected ≥8 headline views, got {headline_count}"

    def test_every_view_has_column_definitions(self):
        for v in m.VIEWS:
            assert v.columns, f"View {v.name} has no column definitions"

    def test_no_sql_injection_risk_in_view_names(self):
        for v in m.VIEWS:
            assert not any(c in v.name for c in " ;'\"`*"), (
                f"View {v.name} has unsafe chars"
            )


# --------------------------------------------------------------------------
# Each view executes against the real warehouse
# --------------------------------------------------------------------------

class TestViewsExecute:

    def test_all_views_run_without_error(self, real_warehouse):
        wh_path, _ = real_warehouse
        con = duckdb.connect(str(wh_path), read_only=True)
        results = m.run_all_views(con)
        con.close()
        failures = [r for r in results if r.error is not None]
        assert not failures, (
            "Views failed:\n" + "\n".join(f"  {r.view.name}: {r.error}" for r in failures)
        )

    def test_all_headline_views_return_rows(self, real_warehouse):
        """Headline views must return rows IF they have a positive expected_min_rows.
        Science-portfolio views (organisms, methods, databases, etc.) declare
        expected_min_rows=0 because they're populated by --extract; they
        legitimately return 0 rows on a Phase-1-only warehouse."""
        wh_path, _ = real_warehouse
        con = duckdb.connect(str(wh_path), read_only=True)
        results = m.run_all_views(con)
        con.close()
        for r in results:
            if r.view.is_headline and r.error is None and r.view.expected_min_rows > 0:
                assert r.row_count >= 1, (
                    f"Headline view {r.view.name} returned 0 rows — ship-blocker"
                )

    def test_all_sanity_floors_met(self, real_warehouse):
        wh_path, _ = real_warehouse
        con = duckdb.connect(str(wh_path), read_only=True)
        results = m.run_all_views(con)
        con.close()
        warnings = [r for r in results if r.error is None and not r.sanity_passed]
        if warnings:
            msg = "\n".join(
                f"  {r.view.name}: got {r.row_count}, expected ≥{r.view.expected_min_rows}"
                for r in warnings
            )
            pytest.fail(f"Sanity-floor violations:\n{msg}")


# --------------------------------------------------------------------------
# Specific view content checks — the headline metrics MUST agree with
# Phase 0 reconnaissance findings
# --------------------------------------------------------------------------

class TestHeadlineContent:

    def test_corpus_summary_matches_phase_0(self, real_warehouse):
        wh_path, _ = real_warehouse
        con = duckdb.connect(str(wh_path), read_only=True)
        df = con.execute([v for v in m.VIEWS if v.name == "corpus_summary"][0].sql).df()
        con.close()
        assert len(df) == 1
        row = df.iloc[0]
        # Corpus size bumped from 53 → 54 on 2026-04-19 with cherry-pick
        # of projects/genotype_to_phenotype_enigma from upstream branch.
        # Using >=53 to stay forward-compatible with future corpus growth.
        assert row["projects"] >= 53
        assert row["notebooks"] >= 241
        assert row["distinct_orcids"] >= 7
        assert row["distinct_reuse_pairs"] >= 75

    def test_top_cited_is_conservation_vs_fitness(self, real_warehouse):
        wh_path, _ = real_warehouse
        con = duckdb.connect(str(wh_path), read_only=True)
        df = con.execute([v for v in m.VIEWS if v.name == "top_cited_projects"][0].sql).df()
        con.close()
        top = df.iloc[0]
        assert top["dst_project_id"] == "conservation_vs_fitness"
        assert top["in_degree"] >= 10  # Phase 0 saw 16

    def test_deepest_revision_is_functional_dark_matter(self, real_warehouse):
        wh_path, _ = real_warehouse
        con = duckdb.connect(str(wh_path), read_only=True)
        view = [v for v in m.VIEWS if v.name == "project_inventory"][0]
        df = con.execute(view.sql).df()
        con.close()
        # Highest revision_depth should be functional_dark_matter
        with_depth = df.dropna(subset=["revision_depth"]).sort_values(
            "revision_depth", ascending=False)
        assert with_depth.iloc[0]["project_id"] == "functional_dark_matter"

    def test_amplification_rate_matches_phase_0(self, real_warehouse):
        wh_path, _ = real_warehouse
        con = duckdb.connect(str(wh_path), read_only=True)
        df = con.execute([v for v in m.VIEWS if v.name == "amplification_rate"][0].sql).df()
        con.close()
        row = df.iloc[0]
        # Phase 0: 33/53 = 62.3%. 2026-04-19 cherry-pick bumped to 54+.
        assert row["total_projects"] >= 53
        assert row["projects_with_declared_citation"] >= 30


# --------------------------------------------------------------------------
# Exporters — CSV + XLSX
# --------------------------------------------------------------------------

class TestExporters:

    def test_full_export_pipeline(self, real_warehouse, tmp_path):
        wh_path, _ = real_warehouse
        rc = atlas_metrics.main([
            "--warehouse", str(wh_path),
            "--outputs", str(tmp_path),
            "--quiet",
        ])
        assert rc == 0, "Metrics exporter exited non-zero"

        metrics_dir = tmp_path / "metrics"
        assert metrics_dir.exists()

        # CSVs
        csv_dir = metrics_dir / "csv"
        assert csv_dir.exists()
        csv_files = list(csv_dir.glob("*.csv"))
        # Every registered view should have produced a CSV (unless it errored,
        # in which case test_all_views_run_without_error would have failed)
        assert len(csv_files) == len(m.VIEWS), (
            f"Expected {len(m.VIEWS)} CSVs, got {len(csv_files)}"
        )

        # XLSX
        xlsx = metrics_dir / "atlas_metrics.xlsx"
        assert xlsx.exists()
        # Open to sanity-check structure
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, read_only=False)
        # TOC + provenance + one sheet per view
        expected_sheets = {"TOC", "__provenance__"} | {v.name[:31] for v in m.VIEWS}
        assert expected_sheets.issubset(set(wb.sheetnames)), (
            f"Missing sheets: {expected_sheets - set(wb.sheetnames)}"
        )

        # Summary JSON
        import json
        summary = json.loads((metrics_dir / "run_summary.json").read_text())
        assert summary["views_failed"] == 0
        assert summary["views_succeeded"] == len(m.VIEWS)

    def test_top_cited_csv_has_expected_shape(self, real_warehouse, tmp_path):
        wh_path, _ = real_warehouse
        atlas_metrics.main([
            "--warehouse", str(wh_path),
            "--outputs", str(tmp_path),
            "--quiet",
        ])
        csv = tmp_path / "metrics" / "csv" / "top_cited_projects.csv"
        assert csv.exists()
        import pandas as pd
        df = pd.read_csv(csv)
        assert "dst_project_id" in df.columns
        assert "in_degree" in df.columns
        # Top row is conservation_vs_fitness
        assert df.iloc[0]["dst_project_id"] == "conservation_vs_fitness"


# --------------------------------------------------------------------------
# Provenance correctness
# --------------------------------------------------------------------------

class TestProvenance:

    def test_provenance_df_matches_view_columns(self):
        for v in m.VIEWS:
            prov_df = m.provenance_dataframe(v)
            assert len(prov_df) == len(v.columns)
            assert set(prov_df["column_name"]) == set(v.columns.keys())

    def test_all_view_columns_documented(self, real_warehouse):
        """Every column returned by SQL must have a definition in view.columns."""
        wh_path, _ = real_warehouse
        con = duckdb.connect(str(wh_path), read_only=True)
        for v in m.VIEWS:
            try:
                df = con.execute(v.sql).df()
            except Exception:
                continue
            sql_cols = set(df.columns)
            defined_cols = set(v.columns.keys())
            undocumented = sql_cols - defined_cols
            assert not undocumented, (
                f"View {v.name} returns columns without provenance: {undocumented}"
            )
        con.close()
